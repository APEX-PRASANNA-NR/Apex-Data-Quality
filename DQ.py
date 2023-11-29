# Import the required packages and modules

#Connector Packages:
from snowflake.snowpark.session import Session
from pyspark.sql import functions as F
import openpyxl as xl

#ML Packages:
from config import *
import numpy as np
import pandas as pd

from scipy.stats import norm
from datetime import datetime

#Other Packages:
import os

db_name = input("DATABASE NAME : ").upper()
schema_name = input("SCHEMA NAME : ").upper()
summary = {}
excel_columns = ["COLUMN_NAME",]

un = input("Username [PRASANNARAVIKUMAR]: ")
if(un == ""):
    un = "PRASANNARAVIKUMAR"

pwd = input("Password [default] : ")
if(pwd == ""):
    pwd = "241020@Nr"

workbook = xl.Workbook()
def snowpark_session():
  conn_params = {
      "account" : "pr91731-production_northeurope",
      "user" : un,
      "password" : pwd,
      "database" : db_name,     
      "role" : "DATAENGINEER",
      "warehouse" : "INGESTION_WH"
  }

  session = Session.builder.configs(conn_params).create()

  return session
try:
    snowpark = snowpark_session()
    print("Connected Successfully")
except Exception as e:
    print(e)

# Connect with the Snowflake to read the data

def write_to_excel_file(worksheet,summary):
    row = 2
    
    #Adding heading
    for excel_column_name in excel_columns:
        # print(excel_column_name,end=" ")
        worksheet.cell(row = 1, column = excel_columns.index(excel_column_name)+1, value = excel_column_name)
    # print()
    for column in summary.keys():
        # print(column,end=" > ")
        worksheet.cell(
            row = row, 
            column = 1,
            value = column.lower()
        )
        
        calculations = summary[column]
        
        for metric in calculations.keys():
            column = excel_columns.index(metric) + 1
            # print("\n\t",column," - ",calculations[metric],end="")
            if(isinstance(calculations[metric],str)):
                cal = calculations[metric].lower()
            else:
                cal = calculations[metric]
            
            worksheet.cell(
                row = row, 
                column = column, 
                value = cal
            )
        row += 1
        # print()
    return worksheet

def show(report,indent = 0,start_delimiter = "[",end_delimiter = "]",seperator = " -> "):    
    if(isinstance(report,list)):
        val = ""
        val += start_delimiter
        for value in report:
            if(report.index(value) == len(report)-1):
                val += str(value)
            else:
                val += str(value)+", "
        val+= end_delimiter
        print(indent+val)
    
    elif(isinstance(report,str)):
        print('"'+report+'"')
    
    
    elif(isinstance(report,dict)):
        for key,value in report.items():
            space = indent*'\t'
            print(f"{space}{key} {seperator} ",end="")
            if(isinstance(value,dict) or isinstance(value,list) or isinstance(value,tuple) or isinstance(value,set)):
                indent += 1
                print("")
                show(indent = indent,report = value)
                indent -=1
            elif(isinstance(value,str)):
                show(indent = indent,report = value)
            else:
                print(f"{value}")

def read_validity_check_list(table_name):
    try:
        validity_workbook = xl.load_workbook("DQ_Validity.xlsx")
        
        if(table_name in validity_workbook.sheetnames):
            
            worksheet = validity_workbook[table_name]
            
            check_list = {}
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                outer_key = row[0]
                inner_dict = {}
                for col_num, value in enumerate(row[2:], start=3):
                    if value is not None:
                        header_value = worksheet.cell(row=1, column=col_num).value
                        inner_dict[header_value] = value
                check_list[outer_key] = inner_dict
                
            validity_workbook.close()
        return check_list
    except Exception as e:
        print(e)
        return False

def check_validity(table_name,col,col_check_list):
    col_data_sql = f'SELECT "{col}" FROM {db_name}.{schema_name}.{table_name}'    
    col_data = pd.DataFrame(snowpark.sql(col_data_sql).collect())
    print(col," : ",col_check_list)
    for checks,val in col_check_list.items():
        if(checks == "format"):
            if(val == "num"):
                numeric_values_count = pd.to_numeric(col_data[col], errors='coerce').dropna().astype(int).count()
                print(summary[col]['TOTAL_COUNT'],summary[col]['NULL_COUNT'],numeric_values_count)
            if(val == "string"):
                if('expected_pattern' in col_check_list.keys()):
                    print(col_data[col],col_check_list[col]['expected_pattern'])
                else:
                    print(col_data[col])

        print()
def data_quality_check(table_name):
  try:
    
    ddl_sql = f"DESCRIBE TABLE {db_name}.{schema_name}.{table_name}"
    ddl = snowpark.sql(ddl_sql).collect()
    
    #collecting the data structure(Datatype)/schematics of the table for identifying the 
    data_sql = f"SELECT * FROM {db_name}.{schema_name}.{table_name}"
    data = snowpark.sql(data_sql)
    
    for row in ddl:
      
      col = row["name"]
      summary[col] = {}
      
      summary[col]["DATA_TYPE"] = row["type"]
      if("DATA_TYPE" not in excel_columns):
        excel_columns.append("DATA_TYPE")
      
      #FINDING THE NULL COUNT IN EACH COLUMN
      summary[col]["NULL_COUNT"] = data.filter(data['"'+col+'"'].isNull()).count()
      if("NULL_COUNT" not in excel_columns):
        excel_columns.append("NULL_COUNT")

      #FINDING THE TOTAL COUNT IN EACH COLUMN
      summary[col]["TOTAL_COUNT"] = data.count()
      if("TOTAL_COUNT" not in excel_columns):
        excel_columns.append("TOTAL_COUNT")
      
      #FINDING THE NOT-NULL COUNT IN EACH COLUMN
      summary[col]["NOT_NULL_COUNT"] = summary[col]["TOTAL_COUNT"] - summary[col]["NULL_COUNT"]
      if("NOT_NULL_COUNT" not in excel_columns):
        excel_columns.append("NOT_NULL_COUNT")
      
      #FINDING THE UNIQUE(DISTINCT) VALUE COUNT IN EACH COLUMN
      summary[col]["UNIQUE_COUNT"] = data.select('"'+col+'"').distinct().count() - 1
      if("UNIQUE_COUNT" not in excel_columns):
        excel_columns.append("UNIQUE_COUNT")
        
      #FINDING THE REPEARING/DUPLICATE VALUE COUNT IN EACH COLUMN EXCLUDING THE UNIQUE COUNT
      summary[col]["DUPLICATE"] = summary[col]["NOT_NULL_COUNT"] - summary[col]["UNIQUE_COUNT"]
      if("DUPLICATE" not in excel_columns):
        excel_columns.append("DUPLICATE")
      
      #Completeness
      summary[col]["COMPLETENESS"] ="{:.2f}".format((summary[col]["NOT_NULL_COUNT"] / summary[col]["TOTAL_COUNT"])*100)
      if "COMPLETENESS" not in excel_columns:
        excel_columns.append("COMPLETENESS")
        
      summary[col]["UNIQUENESS"] ="{:.2f}".format((summary[col]["UNIQUE_COUNT"] / summary[col]["TOTAL_COUNT"])*100)
      if "UNIQUENESS" not in excel_columns:
        excel_columns.append("UNIQUENESS")
      
      
      summary[col]["ACCURACY"] ="{:.2f}".format((summary[col]["NULL_COUNT"] / summary[col]["TOTAL_COUNT"])*100)
      if "ACCURACY" not in excel_columns:
        excel_columns.append("ACCURACY")
    
      check_list = read_validity_check_list(table_name)
    #   print(data['"'+col+'"']," : ",check_list)      
      
      if(check_list):
        col_check_list = check_list[col.lower()]
        
        check_validity(table_name,col,col_check_list)

      else:
        summary[col]["Error"] = "Cannot perform validity as checklist not available"
        if("Error" not in excel_columns):
          excel_columns.append("Error")
      
  except Exception as e:
    print("Error in calculating common data types - ",e)
    
  return summary

all_tables = snowpark.sql(f"SHOW TABLES IN {db_name}.{schema_name}").collect()
table_names = [row["name"] for row in all_tables]

download_folder_path = os.path.expanduser("~" + os.path.sep + "Downloads")
report_path = os.path.join(download_folder_path, 'Data_Quality_Report.xlsx')

for table_name in table_names:
    print(f"Table selected : {table_name}")
    
    if(table_name == 'COHORT_STAGING_MASTER_DATA'):
        
        if(len(workbook.sheetnames) == 1):
            new_worksheet = workbook.active
            new_worksheet.title = table_name
        else:
            new_worksheet = workbook.create_sheet(title = table_name)
        
        report_summary = data_quality_check(table_name)
        # show(report = report_summary)    
        new_worksheet = write_to_excel_file(new_worksheet,report_summary,)
    
downloads_folder = os.path.expanduser("~" + os.path.sep + "Downloads")
file_path = os.path.join(downloads_folder,'Data_Quality_Report_'+schema_name+"("+str(datetime.now().strftime("%d-%m-%Y"))+').xlsx')
workbook.save(file_path)
print("File created...")
#Validation

workbook = xl.load_workbook(file_path)

# Get the number of sheets
number_of_sheets = len(workbook.sheetnames)

snowpark.sql("USE {db_name};")
snowflake_tables = list(snowpark.sql("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = '{schema_name}';").collect())

excel_sheets = list(workbook.sheetnames)

print(len(snowflake_tables),len(excel_sheets))

print("Worksheet not created for : ",set(snowflake_tables)-set(excel_sheets))
print("Unwanted worksheet :", set(excel_sheets)-set(snowflake_tables))